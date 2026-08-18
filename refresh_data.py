# refresh_data.py
r"""Convert the source workbooks into the slim data files the app reads.

Run locally whenever either source export changes, then commit data/.

    python refresh_data.py
"""

import argparse
import json
from pathlib import Path

import openpyxl
import pandas as pd

from size_import.categories import build_lookup, to_json_dict

DEFAULT_CATALOGUE = Path(r"C:\Users\blank\Downloads\Main catalogue.xlsx")
DEFAULT_CATEGORIES = Path(r"C:\Users\blank\Downloads\Glasses size category ids.xlsx")
DATA_DIR = Path(__file__).parent / "data"

NAME_COLUMN = 2         # column C, zero-based
GLOBAL_ID_COLUMN = 103  # column CZ, zero-based


def refresh_catalogue(source, destination):
    workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    records = []
    skipped = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = row[NAME_COLUMN]
        global_id = row[GLOBAL_ID_COLUMN]
        if global_id is None or name is None:
            skipped += 1
            continue
        records.append((str(name).strip(), int(global_id)))

    frame = pd.DataFrame(records, columns=["name", "globalId"])
    frame.to_parquet(destination, index=False)
    print(f"catalogue: {len(frame)} products -> {destination}")
    print(f"catalogue: {skipped} rows skipped (no name or no globalId)")


def refresh_categories(source, destination):
    workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    rows = [row[:3] for row in sheet.iter_rows(min_row=2, values_only=True) if row[0]]
    lookup, report = build_lookup(rows)

    destination.write_text(
        json.dumps(to_json_dict(lookup), indent=1, sort_keys=True), encoding="utf-8"
    )

    print(f"categories: {len(rows)} rows read")
    print(f"categories: {report['collapsed']} duplicates collapsed (lowest ID kept)")
    print(f"categories: {len(report['dropped'])} rows dropped:")
    for category_id, name, value in report["dropped"]:
        print(f"  - id={category_id} name={name!r} value={value!r}")
    print(f"categories: {report['kept']} usable categories -> {destination}")


def main():
    parser = argparse.ArgumentParser(description="Refresh the slim data files.")
    parser.add_argument("--catalogue", type=Path, default=DEFAULT_CATALOGUE)
    parser.add_argument("--categories", type=Path, default=DEFAULT_CATEGORIES)
    args = parser.parse_args()

    DATA_DIR.mkdir(exist_ok=True)
    refresh_catalogue(args.catalogue, DATA_DIR / "catalogue.parquet")
    refresh_categories(args.categories, DATA_DIR / "categories.json")


if __name__ == "__main__":
    main()

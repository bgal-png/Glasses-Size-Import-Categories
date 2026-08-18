# tests/test_export.py
import datetime
from io import BytesIO

import openpyxl

from size_import.export import build_workbook, export_filename, to_bytes


def test_export_filename_uses_two_digit_date():
    assert export_filename(datetime.date(2026, 8, 18)) == "Sizes-260818-import.xlsx"


def test_workbook_has_no_header_row():
    sheet = build_workbook([(1588262, "4156;4157")]).active
    assert sheet["A1"].value == 1588262
    assert sheet["B1"].value == "4156;4157"
    assert sheet.max_row == 1


def test_workbook_writes_one_row_per_product():
    sheet = build_workbook([(1, "10"), (2, "20;30")]).active
    assert [(row[0].value, row[1].value) for row in sheet.iter_rows()] == [
        (1, "10"),
        (2, "20;30"),
    ]


def test_global_id_is_written_as_a_number():
    sheet = build_workbook([(1588262, "4156")]).active
    assert isinstance(sheet["A1"].value, int)


def test_to_bytes_produces_a_readable_workbook():
    data = to_bytes([(1588262, "4156;4157")])
    sheet = openpyxl.load_workbook(BytesIO(data)).active
    assert sheet["A1"].value == 1588262
